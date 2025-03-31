import requests
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

# 配置常量
JIRA_SEARCH_API = "https://issues.apache.org/jira/rest/api/2/search"
JIRA_ISSUE_DETAIL_API = "https://issues.apache.org/jira/rest/api/2/issue/"
JIRA_BROWSE_URL = "https://issues.apache.org/jira/browse/"
EXCEL_FILE = "apache_memory_bugs.xlsx"
MAX_TOTAL_ISSUES = 50  # 最多抓取多少条
PAGE_SIZE = 50  # 单页抓取bug数量
MIN_LOG_LINE = 100  # bug附带的log最小行数
SAVE_EVERY = 10  # 每写入 N 个 issue 就保存一次 Excel


def fetch_memory_bugs():
    start_at = 0
    all_bugs = []

    while start_at < MAX_TOTAL_ISSUES:
        params = {
            "jql": 'issuetype=Bug AND text~"memory"',
            "startAt": start_at,
            "maxResults": PAGE_SIZE
        }

        response = requests.get(JIRA_SEARCH_API, params=params)
        response.raise_for_status()
        data = response.json()

        issues = data.get("issues", [])
        if not issues:
            break

        all_bugs.extend(issues)
        print(f"Fetched {len(issues)} issues (startAt={start_at})")

        start_at += PAGE_SIZE
        if start_at >= data.get("total", 0):
            break

    return all_bugs


def fetch_attachments_with_linecount(issue_key):
    """返回 issue_key, [(attachment_link, line_count)]"""
    url = f"{JIRA_ISSUE_DETAIL_API}{issue_key}"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        attachments = data.get("fields", {}).get("attachment", [])
        result = []

        for att in attachments:
            att_url = att["content"]
            try:
                content_resp = requests.get(att_url, timeout=10)
                content_resp.raise_for_status()
                text = content_resp.text
                line_count = len(text.strip().splitlines())
            except Exception as e:
                print(f"⚠️ 附件无法获取：{att_url}，Error：{e}")
                line_count = "N/A"

            result.append((att_url, line_count))

        return issue_key, result

    except Exception as e:
        print(f"❌ 获取 issue {issue_key} 附件失败: {e}")
        return issue_key, []


def load_written_issue_keys():
    if not os.path.exists(EXCEL_FILE):
        return set()

    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    keys = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            keys.add(row[0])
    return keys


def save_to_excel_incremental(bugs, attachment_map, save_every=SAVE_EVERY):
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Memory Bugs"
        sheet.append(["Issue Key", "Summary", "Issue Link", "Attachment Link", "Attachment type & lines"])

    written_keys = load_written_issue_keys()
    count_since_last_save = 0

    print("\n📄 增量写入 Excel 文件（每写入 {} 条自动保存）...\n".format(save_every))

    for bug in tqdm(bugs, desc="写入进度", ncols=100):
        key = bug.get("key")
        if key in written_keys:
            continue

        summary = bug.get("fields", {}).get("summary", "")
        issue_link = f"{JIRA_BROWSE_URL}{key}"
        attachments = attachment_map.get(key, [])

        if attachments:
            excel_line = [key, summary, issue_link]
            for attachment_link, line_count in attachments:
                # filter logs < 100 lines
                if isinstance(line_count, int) and line_count < MIN_LOG_LINE:
                    continue
                file_type = Path(attachment_link).suffix[1:] or "unknown"
                excel_line.append(attachment_link)
                excel_line.append(f"{file_type}, {line_count}")
            sheet.append(excel_line)
        else:
            sheet.append([key, summary, issue_link, "None", "0"])

        written_keys.add(key)
        count_since_last_save += 1

        if count_since_last_save >= save_every:
            workbook.save(EXCEL_FILE)
            print(f"💾 中间保存 Excel（已写入 {len(written_keys)} 个 issue）")
            count_since_last_save = 0

    workbook.save(EXCEL_FILE)
    print(f"\n✅ 最终写入完成：{EXCEL_FILE}（共写入 {len(written_keys)} 个 issue）")


def main():
    bugs = fetch_memory_bugs()
    print(f"\n📦 共获取到 {len(bugs)} 个 memory 相关的 Bug。\n")

    attachment_map = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_key = {executor.submit(fetch_attachments_with_linecount, bug["key"]): bug["key"] for bug in bugs}
        for future in as_completed(future_to_key):
            try:
                issue_key, attachments = future.result()
                attachment_map[issue_key] = attachments
            except Exception as e:
                print(f"⚠️ 获取附件失败：{e}")

    save_to_excel_incremental(bugs, attachment_map)


if __name__ == "__main__":
    main()
