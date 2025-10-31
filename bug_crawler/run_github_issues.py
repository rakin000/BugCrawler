import os
import re
import json
import requests
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from service.gpt_service.util import get_gpt_answer

def load_config(file):
    with open(file, "r", encoding="utf-8") as f:
        return json.load(f)

config_file = "config/memory_bug/config.json"
config = load_config(config_file)

# GitHub config (falls back to jira keys when appropriate)
GITHUB_SEARCH_API = config.get("github", {}).get("search_api", "https://api.github.com/search/issues")
GITHUB_TOKEN = config.get("github", {}).get("token")
GITHUB_SEARCH_QUERY = config.get("github", {}).get("search_query", config.get("jira", {}).get("jql", ""))
BUG_TYPE = config.get("github", {}).get("bug_type", config.get("jira", {}).get("bug_type"))
MAX_TOTAL_ISSUES = config.get("github", {}).get("max_total_issues", config.get("jira", {}).get("max_total_issues", 500))
PAGE_SIZE = config.get("github", {}).get("page_size", config.get("jira", {}).get("page_size", 50))
MIN_LOG_LINE = config.get("github", {}).get("min_log_line", config.get("jira", {}).get("min_log_line", 10))
SAVE_EVERY = config.get("github", {}).get("save_every", config.get("jira", {}).get("save_every", 10))
ATTACHMENT_FILE_TYPES = config.get("github", {}).get("attachment_file_types", config.get("jira", {}).get("attachment_file_types", []))
GPT_MAX_ATTACHMENT_LINE = config.get("github", {}).get("gpt_max_attachment_line", config.get("jira", {}).get("gpt_max_attachment_line", 2000))
LOG_SAVE_PATH = config.get("github", {}).get("log_save_path", config.get("jira", {}).get("log_save_path", "bug_cases/logs/"))

# excel config
EXCEL_FILE = config["excel"]["file_name"].format(bug_type=BUG_TYPE)

# predefined rules & prompt question
PREDEFINED_RULE_FILE = './prompt_template/predefined_rules.txt'
with open(PREDEFINED_RULE_FILE, 'r') as file:
    PREDEFINED_RULE_FOR_GPT = file.read()
PROMPT_QUESTION_FILE = [
    './prompt_template/question_reason_process_relationship.txt',
    './prompt_template/question_calculate_process_memory_usage.txt'
    # './prompt_template/question_deduce_bug_root_cause.txt'
]
QUESTIONS_FOR_GPT = [open(prompt_template_file, 'r').read() for prompt_template_file in PROMPT_QUESTION_FILE]

# helper to build auth headers for GitHub
HEADERS = {
    "Accept": "application/vnd.github.v3+json"
}
if GITHUB_TOKEN:
    HEADERS["Authorization"] = f"token {GITHUB_TOKEN}"

def fetch_memory_bugs():
    """Search GitHub issues using the configured query and return list of issue items.

    Returns a list of GitHub issue JSON objects from the search API.
    """
    query = GITHUB_SEARCH_QUERY or BUG_TYPE or ""
    if not query:
        raise ValueError("No GitHub search query or BUG_TYPE provided in config")

    per_page = PAGE_SIZE
    page = 1
    all_bugs = []

    while len(all_bugs) < MAX_TOTAL_ISSUES:
        params = {
            "q": query,
            "per_page": per_page,
            "page": page
        }
        response = requests.get(GITHUB_SEARCH_API, params=params, headers=HEADERS, timeout=15)
        response.raise_for_status()
        data = response.json()

        items = data.get("items", [])
        if not items:
            break

        all_bugs.extend(items)
        print(f"Fetched {len(items)} issues (page={page})")

        # GitHub search provides total_count
        if len(all_bugs) >= min(data.get("total_count", 0), MAX_TOTAL_ISSUES):
            break

        page += 1

    return all_bugs[:MAX_TOTAL_ISSUES]


def _extract_urls_from_text(text):
    if not text:
        return []
    # simple url extractor
    urls = re.findall(r"(https?://[^\s)\]\"]+)", text)
    return urls


def fetch_attachments_with_linecount(issue_item):
    """Given a GitHub issue search item (from search API), return (issue_key, [(attachment_link, file_name, line_count)])

    We consider links in the issue body and comments as attachments. Filters by extension using ATTACHMENT_FILE_TYPES when provided.
    """
    try:
        issue_api_url = issue_item.get("url")  # api URL for issue
        repo_api_url = issue_item.get("repository_url")
        number = issue_item.get("number")
        repo_full = ""
        if repo_api_url:
            # repo_api_url looks like https://api.github.com/repos/owner/repo
            repo_full = "/".join(repo_api_url.split("/")[-2:])
        issue_key = f"{repo_full}#{number}" if repo_full else str(number)

        # fetch issue details (to get body) and comments
        result = []

        # Fetch issue details (body)
        if issue_api_url:
            resp = requests.get(issue_api_url, headers=HEADERS, timeout=10)
            resp.raise_for_status()
            issue_detail = resp.json()
        else:
            issue_detail = issue_item

        body = issue_detail.get("body", "")
        urls = set(_extract_urls_from_text(body))

        # fetch comments
        comments_url = issue_detail.get("comments_url")
        if comments_url:
            try:
                c_resp = requests.get(comments_url, headers=HEADERS, timeout=10)
                c_resp.raise_for_status()
                comments = c_resp.json()
                for c in comments:
                    urls.update(_extract_urls_from_text(c.get("body", "")))
            except Exception as e:
                print(f"⚠️ Couldn't fetch comments for {issue_key}: {e}")

        # Filter and fetch line counts
        for url in sorted(urls):
            file_name = Path(url).name.split("?")[0]
            ext = Path(file_name).suffix.lstrip('.').lower()
            # If ATTACHMENT_FILE_TYPES is non-empty, prefer those; else accept any link
            if ATTACHMENT_FILE_TYPES and ext and ext not in ATTACHMENT_FILE_TYPES:
                # skip non-interesting file types
                continue

            try:
                content_resp = requests.get(url, headers=HEADERS, timeout=15)
                content_resp.raise_for_status()
                text = content_resp.text
                line_count = len(text.strip().splitlines())
            except Exception as e:
                print(f"⚠️ 附件无法获取：{url}，Error：{e}")
                line_count = "N/A"

            result.append((url, file_name or url, line_count))

        return issue_key, result

    except Exception as e:
        print(f"❌ 获取 issue 附件失败: {e}")
        return str(issue_item.get("number", "unknown")), []


def load_written_issue_keys():
    if not os.path.exists(EXCEL_FILE):
        return set()

    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    keys = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            keys.add(row[0])
    return keys


def save_to_excel_incremental(bugs, attachment_map, save_every=SAVE_EVERY):
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Memory Bugs"
        sheet.append(["Issue Key", "Summary", "Issue Link", "Attachment Link", "Attachment type & lines", "GPT response"])

    written_keys = load_written_issue_keys()
    count_since_last_save = 0

    print("\n📄 增量写入 Excel 文件（每写入 {} 条自动保存）...\n".format(save_every))

    for bug in tqdm(bugs, desc="写入进度", ncols=100):
        # The search results from GitHub have different shape. Use repository#number as key when possible.
        repo_api = bug.get("repository_url")
        number = bug.get("number")
        if repo_api:
            repo_full = "/".join(repo_api.split("/")[-2:])
            key = f"{repo_full}#{number}"
        else:
            # fallback to html_url
            key = bug.get("html_url") or f"{bug.get('id')}"
        # skip lines that are already written
        if key in written_keys:
            print(f"Case {key} already exists in Excel. Skipping write operation to Excel.")
            continue

        # summary and issue link
        summary = bug.get("title") or bug.get("fields", {}).get("summary", "")
        issue_link = bug.get("html_url") or f"https://github.com/{key}"
        attachments = attachment_map.get(key, [])

        if attachments:
            excel_line = [key, summary, issue_link]
            for index, (attachment_link, attachment_file_name, line_count) in enumerate(attachments):
                # filter logs < 100 lines
                if isinstance(line_count, int) and line_count < MIN_LOG_LINE:
                    continue
                file_type = Path(attachment_link).suffix[1:] or "unknown"
                excel_line.append(attachment_link)
                excel_line.append(f"{file_type}, {line_count}")
                # insert GPT response into the excel result.
                if file_type in ATTACHMENT_FILE_TYPES:
                    # download log file into local folder '/bug_cases/logs'.
                    try:
                        response = requests.get(attachment_link, headers=HEADERS, timeout=20)
                        full_save_path = os.path.join(LOG_SAVE_PATH, key.replace('/', '_')) + f"/{attachment_file_name}"
                        if response.status_code == 200:
                            # Ensure the directory exists
                            os.makedirs(os.path.dirname(full_save_path), exist_ok=True)
                            with open(full_save_path, 'wb') as f:
                                f.write(response.content)
                            print(f"{key} - {attachment_file_name} is downloaded successfully.")
                        else:
                            print(f"⚠️ 下载失败：{attachment_link} 状态码 {response.status_code}")
                    except Exception as e:
                        print(f"⚠️ 下载附件出错 {attachment_link}: {e}")
                    # OpenAI restrict GPT-4 model's maximum context length is 8192 tokens.
                    if line_count > GPT_MAX_ATTACHMENT_LINE:
                        continue
                    for q_index, question in enumerate(QUESTIONS_FOR_GPT):
                        if q_index == 0:
                            question = PREDEFINED_RULE_FOR_GPT + question
                        try:
                            response = get_gpt_answer(question, full_save_path)
                        except Exception as e:
                            response = f"Can't get response from GPT: {e}"
                        excel_line.append(response)
            sheet.append(excel_line)
        else:
            sheet.append([key, summary, issue_link, "None", "0"])

        written_keys.add(key)
        count_since_last_save += 1

        if count_since_last_save >= save_every:
            workbook.save(EXCEL_FILE)
            print(f"💾 中间保存 Excel（已写入 {len(written_keys)} 个 issue）")
            count_since_last_save = 0

    workbook.save(EXCEL_FILE)
    print(f"\n✅ 最终写入完成：{EXCEL_FILE}（共写入 {len(written_keys)} 个 issue）")


def main():
    bugs = fetch_memory_bugs()
    print(f"\n📦 共获取到 {len(bugs)} 个 memory 相关的 Bug。\n")

    attachment_map = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_key = {executor.submit(fetch_attachments_with_linecount, bug["key"]): bug["key"] for bug in bugs}
        for future in as_completed(future_to_key):
            try:
                issue_key, attachments = future.result()
                attachment_map[issue_key] = attachments
            except Exception as e:
                print(f"⚠️ 获取附件失败：{e}")
    save_to_excel_incremental(bugs, attachment_map)


if __name__ == "__main__":
    main()
